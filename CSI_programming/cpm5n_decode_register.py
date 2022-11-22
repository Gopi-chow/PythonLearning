#!/tools/batonroot/rodin/engkits/lnx64/python-3.8.3_slv/bin/python
"""
Script for CPM5N_CDX_CSI_CFG_CSR registers
"""
import argparse
import os
import time
from openpyxl import Workbook
import openpyxl
import time
from itertools import chain
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, NamedStyle

from cpm5n_register import *
from cpm5n_register_utility import *

parser = argparse.ArgumentParser()
parser.add_argument("input", type=str, help="Input file path")
parser.add_argument("-output", default="output.txt", help="Output filename")
parser.add_argument(
    "-decoder",
    choices=["cdo_to_text", "text_to_cdo", "c_to_text"],
    default="cdo_to_text",
    help="Decode utility type:: i.e cdo_to_text or text_to_cdo",
)
parser.add_argument(
    "-add",
    default="0xe4",
    help="Register prefix values for psx/pcie bridge",
)
parser.add_argument("-path", default=os.getcwd(), help="file path to write output file")
args = parser.parse_args()
time_stamp = time.strftime("%Y%m%d-%H%M%S")
file_name = "csi_cfg_csr_" + time_stamp + "_" + args.output
excel_file_name = "csi_cfg_csr_" + time_stamp + "_" + ".xlsx"

reg_values = {"START_ADD_REG": ["ADDR", "SINK ID(RO)", "threshold"],
              "END_ADD_REG": ["ADDR", "SINK ID(RO)"],
              "WR_PTR_REG": ["WR_ADDR", "sink ID (RO)"],
              "RD_PTR_REG": ["RD_ADDR", "sink ID (RO)"],
              "JOB_SIZE_REG": ["prog JS", "curr JS", "curr jon segments"],
              "TMR_VAL_REG": ["Timer value"],
              "TMR_SNST_REG": ["txn_on", "tmr_exp", "tmr_seg", "tmr_armed", "tmr_snp_val"],
              "GBL_BUF_ID": ["glb buf id", "glb buf sink id"],
              "SEQ_CNT_REG": ["PR seq count ID", "NPR?CMPL", "local credit counter"],
              "CAP_MAP_REG": ["CSI Src ID", "CSI virt ID", "Csched dest ID", "SVC managed by cSched", "is_npr",
                              "ctxt_prog_done"],
              "RSLT_NPR_SIZE": ["resultant npr size"],
              "JOB_LEN_AOFF": ["job length buf start addr", "job len buff end addr"],
              "JOB_LEN_PTR": ["job_buf_wr_ptr", "job_buf_rd_ptr"],
              "CAP_MAP_REG_": ["CSI Src ID", "CSI virt ID", "Csched dest ID", "SVC managed by cSched", "is_npr",
                               "ctxt_prog_done"],
              }
excel_file_name = os.path.join(args.path, excel_file_name)
excel_sheet = excel_file_name


def decode_dict(keyword, reg_number, reg_values_log):
    all_reg_values = []
    for key in list(reg_values.keys()):
        key_to_search = keyword + "_" + key + str(reg_number)
        print(key_to_search, reg_values_log.get(key_to_search))
        all_reg_values.append(reg_values_log.get(key_to_search, ["NA"] * len(reg_values[key])))
    return all_reg_values


thick_border = Border(left=Side(style='medium'),
                      right=Side(style='medium'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))


def make_excel_table():
    """To create table with register value and REG."""

    d = ["REG", "FEILD", "REG0", "REG1", "REG2", "REG3", "REG4", "REG5", "REG6", "REG7", "REG8", "REG9", "REG10",
         "REG11"]
    keys = list(reg_values.keys())
    book = Workbook()
    sheet = book.active
    start = 2
    start_row = 2
    types = ["pcie_snk_npr", "pcie_snk_cmpl", "pcie_snk_pr", "cdm_snk_cmpl", "psx_snk_npr", "psx_snk_cmpl",
             "psx_snk_pr"]
    for i in range(7):
        row = start - 1
        for ind, val in enumerate(d):
            sheet.cell(row=row, column=ind + 1).value = val
            sheet.cell(row=row, column=ind + 1).font = Font(bold=True)
            sheet.cell(row=row, column=ind + 1).border = thick_border

        for key in reg_values:
            sheet.cell(row=start, column=1).value = types[i] + "_" + key
            sheet.cell(row=start, column=1).font = Font(bold=True)
            sheet.cell(row=start, column=1).border = thick_border
            row_length = len(reg_values[key])
            end = start + (row_length - 1)
            sheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            start = end + 1

        start_row = row + 1
        for k in reg_values:
            for val in reg_values[k]:
                sheet.cell(row=start_row, column=2).value = val
                sheet.cell(row=start_row, column=2).font = Font(bold=True)
                sheet.cell(row=start_row, column=2).border = thick_border
                start_row += 1

        start = start + 5

    book.save(excel_sheet)


def write_data_into_excel(col, row, values):
    wb_obj = openpyxl.load_workbook(excel_sheet)
    sheet_obj = wb_obj.active
    row = row
    col = col
    for i in values:
        sheet_obj.cell(row=row, column=col).value = i
        sheet_obj.cell(row=row, column=col).border = thick_border
        row += 1
    wb_obj.save(excel_sheet)


def csi_cfg_csr_decode():
    """Decode CPM5N_CDX_CSI_CFG_CSR register"""
    cdo_file = convert_c_file_to_cdo(args.input) if args.decoder == "c_to_text" else args.input
    try:
        file_obj = open(cdo_file, "r")
    except FileNotFoundError as e:
        raise FileNotFoundError(f"File not found {e}")
    file_name = "csi_cfg_csr_" + time_stamp + "_" + args.output
    excel_file_name = "csi_cfg_csr_" + time_stamp + "_" + ".xlsx"
    file_name = os.path.join(args.path, file_name)
    excel_file_name = os.path.join(args.path, excel_file_name)
    if args.decoder == "text_to_cdo":
        csi_cfg_csr_text_to_cdo(file_obj, file_name)
        return
    output = open(file_name, "w")
    count = 0
    add = args.add.lower()
    reg0_d = {}
    count_cap_map_reg = 1
    for f in file_obj:
        count = count + 1
        reg0, reg1, reg2, reg3, reg4, reg5, reg6, reg7, reg8, reg9, reg10, reg11 = [], [], [], [], [], [], [], [], [], [], [], []
        # Searching for CSCHED reg address
        a = re.search(rf"write_print {add}46(.*)$", f)
        if a:
            f_data = a.group(0).strip()
            print(f_data)
            sep = "*" * 50 + "\n"
            d = f_data.replace("write_print ", "").replace("\t", " ")
            print("debug:" + d)
            # Finding offset, Substracting address value from CSCHED base address
            offset = hex(int(d[0:10], 16) - (int(f"{add}460000", 16))).upper()
            f_data = "Raw Value:\n" + str(count) + "-->" + a.group(0).strip()
            if int(offset, 16) >= 0:
                print("The value of d is:")
                print(len(d[10:]))
                print(type(d[10:]))
                print(d)
                print(d[10:])
                t = convert_hex_binary(d[10:])
                if offset[2:] in [a.value for a in START_ADD_REG]:
                    reg = START_ADD_REG(offset[2:]).name
                    print("debug data: t is" + t)
                    print(len(t))
                    s = pcie_snk_npr_START_ADD_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in END_ADD_REG]:
                    reg = END_ADD_REG(offset[2:]).name
                    s = pcie_snk_npr_END_ADD_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in WR_PTR_REG]:
                    reg = WR_PTR_REG(offset[2:]).name
                    s = pcie_snk_npr_WR_PTR_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in RD_PTR_REG]:
                    reg = RD_PTR_REG(offset[2:]).name
                    s = pcie_snk_npr_RD_PTR_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in JOB_SIZE_REG]:
                    reg = JOB_SIZE_REG(offset[2:]).name
                    s = pcie_snk_npr_JOB_SIZE_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in TMR_VAL_REG]:
                    reg = TMR_VAL_REG(offset[2:]).name
                    s = pcie_snk_npr_TMR_VAL_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in TMR_SNST_REG]:
                    reg = TMR_SNST_REG(offset[2:]).name
                    s = pcie_snk_npr_TMR_SNST_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in GBL_BUF_ID]:
                    reg = GBL_BUF_ID(offset[2:]).name
                    s = pcie_snk_npr_GBL_BUF_ID0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in SEQ_CNT_REG]:
                    reg = SEQ_CNT_REG(offset[2:]).name
                    s = pcie_snk_npr_SEQ_CNT_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif (count_cap_map_reg % 2 == 1) and offset[2:] in [a.value for a in CAP_MAP_REG]:
                    reg = CAP_MAP_REG(offset[2:]).name
                    s = pcie_snk_npr_CAP_MAP_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                    output.write(f"debug data: observed {count_cap_map_reg} time in {reg}")
                    count_cap_map_reg = count_cap_map_reg + 1
                elif (count_cap_map_reg % 2 == 0) and offset[2:] in [a.value for a in CAP_MAP_REG]:
                    reg = CAP_MAP_REG(offset[2:]).name
                    index = reg.index("REG") + 3
                    reg = reg[:index] + "_" + reg[index:]
                    s = pcie_snk_npr_CAP_MAP_REG0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                    output.write(f"debug data: Observed {count_cap_map_reg} time in {reg}")
                    count_cap_map_reg = count_cap_map_reg + 1
                elif offset[2:] in [a.value for a in RSLT_NPR_SIZE]:
                    reg = RSLT_NPR_SIZE(offset[2:]).name
                    s = pcie_snk_npr_RSLT_NPR_SIZE0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in JOB_LEN_AOFF]:
                    reg = JOB_LEN_AOFF(offset[2:]).name
                    s = pcie_snk_npr_JOB_LEN_AOFF0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in JOB_LEN_PTR]:
                    reg = JOB_LEN_PTR(offset[2:]).name
                    s = pcie_snk_npr_JOB_LEN_PTR0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
                elif offset[2:] in [a.value for a in SINK_BUF_FILL]:
                    reg = SINK_BUF_FILL(offset[2:]).name
                    s = pcie_snk_npr_SINK_BUF_FILL0(t)
                    print("reg value is:", reg)
                    reg0_d[reg] = s[1].split(",")
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                elif offset[2:] in [a.value for a in READ_GOVERNER]:
                    reg = READ_GOVERNER(offset[2:]).name
                    s = pcie_snk_npr_READ_GOVERNER0(t)
                    output.write(f"{f_data}\n{reg}::{s}\n{sep}")
                    output.write(f"data in current iteraion:{reg0_d[reg]}")
        else:
            output.write(f"{f}")
    print("INFO: CPMS5N_CDX_CSI_CFG_CSR details added into " + file_name)
    return reg0_d


def csi_cfg_csr_text_to_cdo(file_obj, output_file):
    output = open(output_file, "w")
    for j in file_obj:
        i = j.strip().split("::")
        i = [x for x in i if x]
        if len(i) == 2:
            offset = CPM5N_CDX_CSI_CFG_CSR_REG[i[0]].value
            t = i[1]
            address = hex(
                int(f"{args.add.lower()}460000", 16) + int(CPM5N_CDX_CSI_CFG_CSR_REG[i[0]].value, 16)
            )
            if offset in [a.value for a in START_ADD_REG]:
                s = pcie_snk_npr_START_ADD_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in END_ADD_REG]:
                s = pcie_snk_npr_END_ADD_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in WR_PTR_REG]:
                s = pcie_snk_npr_WR_PTR_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in RD_PTR_REG]:
                s = pcie_snk_npr_RD_PTR_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in JOB_SIZE_REG]:
                s = pcie_snk_npr_JOB_SIZE_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in TMR_VAL_REG]:
                s = pcie_snk_npr_TMR_VAL_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in TMR_SNST_REG]:
                s = pcie_snk_npr_TMR_SNST_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in GBL_BUF_ID]:
                s = pcie_snk_npr_GBL_BUF_ID0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in SEQ_CNT_REG]:
                s = pcie_snk_npr_SEQ_CNT_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in CAP_MAP_REG]:
                print("Debug data t is", t)
                s = pcie_snk_npr_CAP_MAP_REG0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in RSLT_NPR_SIZE]:
                s = pcie_snk_npr_RSLT_NPR_SIZE0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in JOB_LEN_AOFF]:
                s = pcie_snk_npr_JOB_LEN_AOFF0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in JOB_LEN_PTR]:
                s = pcie_snk_npr_JOB_LEN_PTR0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in SINK_BUF_FILL]:
                s = pcie_snk_npr_SINK_BUF_FILL0(t)
                output.write(f"write {address} {s}\n")
            if offset in [a.value for a in READ_GOVERNER]:
                s = pcie_snk_npr_READ_GOVERNER0(t)
                output.write(f"write {address} {s}\n")
        else:
            if "Raw Value:" in j or "-->" in j or "*" * 50 in j:
                pass
            else:
                output.write(j)
    print(f"INFO: decoded data stored into {output_file}")


time_stamp = time.strftime("%Y%m%d-%H%M%S")
make_excel_table()
reg0_d = csi_cfg_csr_decode()
col = 3
ex = open("check.txt", "w")
for i in range(12):
    print("Excel for iteration:", i)
    all_reg_values = decode_dict("pcie_snk_npr", i, reg0_d)
    values = list(chain(*all_reg_values))
    write_data_into_excel(col, 2, values)
    col += 1

# "cdm_snk_cmpl","psx_snk_pr"

# Attr_type and column number
types = [("pcie_snk_npr", 2), ("pcie_snk_cmpl", 47), ("pcie_snk_pr", 92), ("cdm_snk_cmpl", 137), ("psx_snk_pr", 182),
         ("psx_snk_cmpl", 227), ("psx_snk_pr", 272)]

for attr_type in types:
    col = 3
    for i in range(12):
        print("Excel for iteration:", i)
        all_reg_values = decode_dict(attr_type[0], i, reg0_d)
        values = list(chain(*all_reg_values))
        write_data_into_excel(col, attr_type[1], values)
        col += 1

print(all_reg_values)
print(list(chain(*all_reg_values)))
print(reg0_d)
