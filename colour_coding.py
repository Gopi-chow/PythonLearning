import os
import glob
import csv
from xlsxwriter.workbook import Workbook
import openpyxl
from openpyxl.styles import PatternFill
from xlsxwriter.workbook import Workbook

csvfile = "CPM5N_CDX_CSI_CFG_CSR_log_file_combined.csv"
excel_file = "CPM5N_CDX_CSI_CFG_CSR_log_file_combined.xlsx"
workbook = Workbook(excel_file)
worksheet = workbook.add_worksheet("CDO") # worksheet with csv file name
with open(csvfile, 'r') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            worksheet.write(r, c, col) # write the csv file content into it
workbook.close()

wb_obj = openpyxl.load_workbook(excel_file) #path to the Excel file
fill_cell_red = PatternFill(patternType='solid', fgColor='C64747')
fill_cell_green = PatternFill(patternType='solid', fgColor='35FC03')

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
for i in range(2, max_row + 1):
    #Get the value of each row in a column-3 which is an object
    expected_value_obj = sheet_obj.cell(row=i, column=3)
    reset_value_obj = sheet_obj.cell(row=i, column=4)
    cdo_value_obj = sheet_obj.cell(row=i, column=5)
    # The object has key value which gives the value of the cell
    if (expected_value_obj.value != cdo_value_obj.value):
        if not cdo_value_obj.value:
            cdo_value_obj.value = "CDO couldnot find"
        if reset_value_obj.value != cdo_value_obj.value:
            cdo_value_obj.fill = fill_cell_red
        elif reset_value_obj.value == cdo_value_obj.value:
            cdo_value_obj.fill = fill_cell_green
    else:
        cdo_value_obj.fill = fill_cell_green
wb_obj.save(excel_file)