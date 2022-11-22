import openpyxl
from openpyxl import Workbook
dc = "h10_cpm_es2_mainline_20220728_7837f45_20220826_660b8ba_FAILED_TEST_HANG.csv"
f = open(dc,"r")
csv_lines = f.readlines()
f.close()

key_data= {
    "H10_RC_H10_EP":"h10_rc_h10_ep",
    "H10_EP":"x86_rc_h10_ep",
    "H10_RC":"h10_rc_k7_ep",
    "VPK120_EP":"x86_rc_vpk120_ep",
    "H10_RC_VPK120_EP":"h10_rc_vpk120_ep"
     }
total_no_cols_to_copy = 19
log_path_index = 16

# first_line = csv_lines[1].strip("\n").split(",")
# count = 0
# for line in first_line:
#     a = [True if each_key_word in line else False for each_key_word in list(key_data.values())]
#     if any(a):
#         log_path_index = first_line.index(line)
#     if line == "":
#         break
#     count += 1
# total_no_cols_to_copy = count

### From a CSV file, it will filter the log file column based on the name and create a sheet in the excel and write the data of CSV into an excel sheet
wb = Workbook()
ws = wb.active
work_book_title = list(key_data.keys())
ws.title = work_book_title[0]
count = 0
excel_sheets = wb.sheetnames
for each_sheet in work_book_title:
    if each_sheet in excel_sheets:
        wb.active = wb[each_sheet]
        ws = wb.active
    else:
        ws1 = wb.create_sheet(each_sheet)
        wb.active = wb[each_sheet]
        ws = wb.active
    for each_line in csv_lines:
        each_line_list = each_line.strip("\n").split(",")[0:total_no_cols_to_copy]
        if len(each_line_list) != total_no_cols_to_copy:
            continue
        if key_data[ws.title] in each_line_list[log_path_index]:
            ws.append(each_line_list)
wb.save(filename = "aug_29_mainline.xlsx")